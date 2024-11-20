import re
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import sys
current_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.append(os.path.join(current_dir,'../'))
from util import *
from excel_to_source_files_info import *
import xml.etree.ElementTree as ET
from collections import OrderedDict

# ios翻译源文件文件夹
input_ios_lproj_path = './trans_to_excel/input/ios/'

# android翻译源文件文件夹
input_android_values_path = './trans_to_excel/input/android/'

# web翻译源文件文件夹
input_web_js_path = './trans_to_excel/input/web/'

# 输出excel文件目录
output_excel_path = './trans_to_excel/output/'
    
# 逐行读取ios国际化源文件
def read_trans_file_ios():
    # 多语言翻译的map 例如:{zh-hans: {'trans0001': "中文"}, en: {'trans0001': "english"}}
    lang_trans_map = OrderedDict()
    dir_path = f'{input_ios_lproj_path}'
    for info in Project.allLanguages(project):
        file_path = f'{dir_path }{info.ios_lproj_file_name}/Localizable.strings'
        string_file_map = _read_string_file_ios(file_path)
        _key = info.excel_lang_id
        lang_trans_map[_key] = string_file_map
    # print('trans_map: ', trans_map)
    return auto_fill_trans_key_by_en(lang_trans_map)

#  处理ios key值
def _read_string_file_ios(file_path):
    with open(file_path, 'r+') as file:
        # 单个文件翻译的map 例如:{'trans0001': "中文"}
        trans_lang_map = OrderedDict()
        for line in file:
            # print(f'{line} = {line}')
            if '=' in line:
                trans = re.split(r"\"\s*=\s*\"", line)
                if len(trans) > 1:
                    key = trans[0].replace('"', '')
                    text = trans[1].replace('";\n', '')
                    # 删除转义符 \
                    for escaped_char in Project.getEscapedCharacters(): 
                        text = text.replace(f'\{escaped_char}', f'{escaped_char}')
                    # print(f'{_key} = {_text}')
                    _text = _convert_to_excel_args(text, Platform.ios)
                    _text = _convert_to_excel_imgs(_text)
                    # 去除前后空格
                    _text = _text.strip()
                    trans_lang_map[key] = _text
        return trans_lang_map
    
# 每种语言按英文补全key后的全量翻译map
def auto_fill_trans_key_by_en(trans_map):
    # 英语的id
    en_lang_id = Project.allLanguages(project)[0].excel_lang_id
    en_trans_map = Util.safe_value(trans_map, en_lang_id) # 以英文为准，其他语言缺失的key用英文补齐
    full_lang_map = {en_lang_id: en_trans_map} # 所有语言按英文补全key后的全量翻译map
    for langkey in trans_map.keys():
        per_trans_map = trans_map[langkey]
        if langkey != en_lang_id:
            full_trans_map = OrderedDict()  # 每种语言按英文补全key后的全量翻译map
            for transKey in en_trans_map.keys():
                if transKey in per_trans_map.keys():
                    full_trans_map[transKey] = per_trans_map[transKey]
                else:
                    full_trans_map[transKey] = ''
            full_lang_map[langkey] = full_trans_map

    # print('full_trans_map: ', full_trans_map)
    return full_lang_map

# 逐行读取android国际化源文件
def read_trans_file_android():
    # 多语言翻译的map 例如:{zh-hans: {'trans0001': "中文"}, en: {'trans0001': "english"}}
    trans_map = OrderedDict()
    dir_path = f'{input_android_values_path}'
    for info in Project.allLanguages(project):
        file_path = f'{dir_path }{info.android_values_file_name}/strings.xml'
        trans_lang_map = _read_xml_file_android(file_path)
        _key = info.excel_lang_id
        trans_map[_key] = trans_lang_map
    # print('trans_map: ', trans_map)
    return auto_fill_trans_key_by_en(trans_map)

# 将嵌套的xml节点转换为字符串,还可能嵌套<bold></bold> / <u></u> / <b></b>
def _get_nested_text(element):
    text = element.text
    for child in element: # 判断值是否嵌套
        text = ET.tostring(child).decode()
        text = text
        break 
    return text

#  处理android key值
def _read_xml_file_android(file_path):
    # 单个文件翻译的map 例如:{'trans0001': "中文"}
    trans_lang_map = OrderedDict()
    # xml路径
    xml_path = f'{file_path}/strings.xml'
    # 加载XML文件
    tree = ET.parse(xml_path)
    root = tree.getroot()
    # 遍历根节点的子元素
    for child in root.iter():
        name = child.get('name')
        if name is None:
            continue        
        # 使用element_to_string函数转换
        text = _get_nested_text(child)
        if text is None:
            continue        
        # key 去掉空格
        key = name.replace(' ', '').replace('\n', '').replace('\t', '')
        # 删除转义符 \
        for escaped_char in Project.getEscapedCharacters(): 
            text = text.replace(f'\{escaped_char}', f'{escaped_char}')
        # 将特殊字符 & 转为 &amp  ’
        text = text.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
        text = _convert_to_excel_args(text, Platform.android)
        text = _convert_to_excel_imgs(text)
        trans_lang_map[key] = text
    return trans_lang_map
    
# 将iOS 安卓 web各端参数转换成excel统一格式{1} 
def _convert_to_excel_args(text, platform):
    value = text
    # 查找占位的参数
    if platform == Platform.ios:
        replace_arg_map = {'%d': '%@', '%1$d': '%1$@', '%2$d': '%2$@', '%3$d': '%3$@', '%4$d': '%4$@', '%5$d': '%5$@', '%6$d': '%6$@', '%7$d': '%7$@', '<customText>': '%@'}
        for old_arg, new_arg in replace_arg_map.items():
            if old_arg in text:
                value = text.replace(old_arg, new_arg) # 将%d参数替换成%@
        result = value.split('%@')
        if len(result) >= 2:
            value = ''
            index = 0
            for res in result:
                value = value + res + '{' + f'{index + 1}' + '}'
                last_index = len(result) - 2
                if index == last_index:
                    value = value + result[-1] # 拼接最后一段
                    break
                index += 1

        # 替换%1$@参数
        for args in re.findall(r'%d$@', text):
            excel_args = args.replace('%', '{').replace('$@', '}')
            value = text.replace(args, excel_args)

    if platform == Platform.android:
        replace_arg_map = {'%d': '%s', '%1$d': '%1$s', '%2$d': '%2$s', '%3$d': '%3$s', '%4$d': '%4$s', '%5$d': '%5$s', '%6$d': '%6$s', '%7$d': '%7$s'}
        for old_arg, new_arg in replace_arg_map.items():
            if old_arg in text:
                value = text.replace(old_arg, new_arg) # 将%d参数替换成%s
        result = value.split('%s')
        if len(result) >= 2:
            value = ''
            index = 0
            for res in result:
                value = value + res + '{' + f'{index + 1}' + '}'
                last_index = len(result) - 2
                if index == last_index:
                    value = value + result[-1] # 拼接最后一段
                    break
                index += 1

        # 替换%1$s参数
        for args in re.findall(r'%d$s', value):
            excel_args = args.replace('%', '{').replace('$s', '}')
            value = value.replace(args, excel_args)
    if platform == Platform.web:
        value = text
    return value

# 将iOS 安卓 web各端[img src=xx]参数转换成excel统一格式{img src=xx} 
def _convert_to_excel_imgs(text):
    value = text
    # 替换[img src=xxx]参数
    for result in re.findall(r'[img .*?]', text):
        img_tag = result.replace('[', '{').replace(']', '}')
        value = value.replace(result, img_tag)
    for result in  re.findall(r'<img .*?>', text):
        img_tag = result.replace(result, '{img}')
        value = value.replace(result, img_tag)
    return value
    
# 将ios国际化源文件读出的多语言写入excel文件
def write_trans_to_excel(dir, trans_map, platform):

    dir = dir
    if platform == Platform.ios:
        dir = dir + 'ios'
    elif platform == Platform.android:
        dir = dir + 'android'
    elif platform == Platform.web:
        dir = dir + 'web'

    file_name = "new_trans.xlsx"  # 新文件的名称
    file_path = os.path.join(dir, file_name)  # 合并目录和文件名
    # 创建文件夹
    if not os.path.exists(dir):
        os.makedirs(dir)
    # 创建excel文件
    with open(file_path, 'r+', encoding='utf-8') as file:
        print(f"文件 {file_name} 已创建在目录 {dir}")

    # print(f'write_trans_to_excel {file_name}...')

    # 英语的id
    en_lang_id = Project.allLanguages(project)[0].excel_lang_id
    en_trans_map = Util.safe_value(trans_map, en_lang_id) # 以英文为准，其他语言缺失的key用英文补齐
    # 按英文关键字自动生成key
    _auto_generate_new_key_map = _auto_generate_key_map(en_trans_map)
    
    workbook = Workbook()
    sheet = workbook.active

    sheet.cell(row=1, column=1).value = 'iOS Key:(Don\'t Modify)'
    sheet.cell(row=1, column=2).value = 'Android Key:(Don\'t Modify)'
    sheet.cell(row=1, column=3).value = 'Web Key:(Don\'t Modify)'
     # excel中key对应的列
    excel_key_col_map = OrderedDict({Platform.ios: 1, Platform.android: 2, Platform.web: 3})
    # 翻译文本从第4列开始
    currentColumn = 4
    for lang in Project.allLanguages(project):
        # 第一行描述
        first_row_desc = lang.excel_lang_id + ":" + lang.desc
        sheet.cell(row=1, column=currentColumn).value = first_row_desc
        
        lang_dict = Util.safe_value(trans_map, lang.excel_lang_id)
        if lang_dict is None:
            continue
        # 翻译文本从第2行开始
        currentRow = 2
        for key in lang_dict:
            # 按英文关键字自动生成key
            _auto_generate_key = _auto_generate_new_key_map[key]

            text = Util.safe_value(lang_dict, key)
            if text is None or text == '\"\"':
                continue
            excel_key_col = excel_key_col_map[platform]
            # 第一列key
            sheet.cell(row=currentRow, column=excel_key_col).value = _auto_generate_key
            # 多语言翻译
            sheet.cell(row=currentRow, column=currentColumn).value = text
            # print(f'writing...: row: {currentRow}, column:{currentColumn}, text: {text}')
            currentRow = currentRow + 1
        currentColumn = currentColumn + 1

    # _set_excel_style(sheet)

    workbook.save(file_path)
    workbook.close

def _set_excel_style(sheet):
    # 自定义字体样式
    font = Font(
        name="Calibri",  # 字体
        size=11,         # 字体大小
        color="000000",  # 字体颜色，用16进制rgb表示
        bold=False,       # 是否加粗，True/False
        italic=False,     # 是否斜体，True/False
        strike=None,     # 是否使用删除线，True/False
        underline=None,  # 下划线, 可选'singleAccounting', 'double', 'single', 'doubleAccounting'
    )
    align = Alignment(
        horizontal='left',     # 水平对齐，可选general、left、center、right、fill、justify、centerContinuous、distributed
        vertical='center',        # 垂直对齐， 可选top、center、bottom、justify、distributed
        text_rotation=0,       # 字体旋转，0~180整数
        wrap_text=True,       # 是否自动换行
        shrink_to_fit=False,   # 是否缩小字体填充
        indent=0,              # 缩进值
    )
    
     # 设置excel样式
    for col in range(1, sheet.max_column+1):
        col_name = get_column_letter(col)
        print(f'_set_excel_style {col_name} col')
        sheet.column_dimensions[col_name].width = 40

        for row in range(0, sheet.max_row):
            sheet[col_name][row].font = font
            sheet[col_name][row].alignment = align

# 新旧key映射map [newKey: oldKey]
def _auto_generate_key_map(trans_map):
    _new_old_key_map = OrderedDict()
    for key, text in trans_map.items():
        # 默认先取count个单词拼接key
        _new_key = Util.create_not_repeat_key_in_map(text, _new_old_key_map, 4)
        _new_old_key_map[key] = _new_key 
    return _new_old_key_map 

# map按value去重
def _find_duplicate_values(dict):
    unique_dict = OrderedDict()
    duplicates_dict = OrderedDict()
    for key, value in dict.items():
        if value in unique_dict.values():
            duplicates_dict[key] = value 
        else:
            unique_dict[key] = value 
    return duplicates_dict

# 按英文去重    
def _delete_duplicate_by_en(lang_trans_map):
     # 英语的id
    en_lang_id = Project.allLanguages(project)[0].excel_lang_id
    en_trans_map = Util.safe_value(lang_trans_map, en_lang_id) 
    duplicates_dict = _find_duplicate_values(en_trans_map)
    all_dict = OrderedDict()
    for key, trans_map in lang_trans_map.items():
        no_duplicates_dict = trans_map
        # 移除重复值的key
        for duplicates_key in duplicates_dict.keys(): 
            no_duplicates_dict.pop(duplicates_key)
        all_dict[key] = no_duplicates_dict
    return lang_trans_map


def trans_to_excel():
    print('正在转换...')
    
    # 读取ios国际化源文件
    trans_map_ios = read_trans_file_ios()
    # 按英文去重    
    _delete_duplicates_map = _delete_duplicate_by_en(trans_map_ios)
    # 写入excel文件
    write_trans_to_excel(output_excel_path, _delete_duplicates_map, Platform.ios)

    # 读取android国际化源文件
    # trans_map_android = read_trans_file_android()
    # # 写入excel文件
    # write_trans_to_excel(output_excel_path, trans_map_android, Platform.android)

    # # 读取web国际化源文件
    # trans_map_web = read_trans_file_web()
    # # 写入excel文件
    # write_trans_to_excel(output_excel_path, trans_map_web, Platform.web)
    
    print('翻译源文件转excel成功...')

# 转换翻译的app, 默认新的jblone
project = Project.jblone

# 函数入口 第一个参数是转换语言的app项目名称  eg: python3 trans_to_excel.py partybox / jblone
if __name__ == '__main__':
    if len(sys.argv) > 1: # 传递了参数
        project = sys.argv[1] # 第一个参数是转换语言的app项目名称 partybox / jblone
    
    trans_to_excel()