import re
import os
import openpyxl
import sys
current_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.append(os.path.join(current_dir,'../'))
from util import *
from excel_to_source_files_info import *
from collections import OrderedDict

# 输入ios excel文件文件夹
input_ios_excel_path = './trans_from_excel/input/ios/'
# 输出ios 翻译源文件
output_ios_lproj_path = './trans_from_excel/output/ios/'

# 输入android excel文件文件夹
input_android_excel_path = './trans_from_excel/input/android/'
# 输出android 翻译源文件
output_android_values_path = './trans_from_excel/output/android/'

# 输入web excel文件文件夹
input_web_excel_path = './trans_from_excel/input/web/'
# 输出web 翻译源文件
output_web_js_path = './trans_from_excel/output/web/'

# 读取excel,生成翻译映射关系 例如:{zh_hans: {'trans0001': "中文"}, en: {'trans0001': "english"}}
def read_excel(excen_path, platform):

    # 导入多语言翻译的excel
    wb = openpyxl.load_workbook(excen_path, data_only=True)

    # 获取所有工作表的名称
    sheet_names = wb.sheetnames

    # 多语言翻译的map 例如:{zh_hans: {'trans0001': "中文"}, en: {'trans0001': "english"}}
    all_sheet_trans_map = OrderedDict()
    # 遍历所有工作表
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        per_sheet_trans_map = read_sheet(wb, excen_path, sheet, platform)
        for lang in per_sheet_trans_map:
            per_lang_map = Util.safe_value(per_sheet_trans_map, lang)
            if per_lang_map is None:
                continue
            all_lang_map = Util.safe_value(all_sheet_trans_map, lang)
            if all_lang_map is not None:
                all_lang_map.update(per_lang_map)    
                all_sheet_trans_map[lang] = all_lang_map
            else:
                all_sheet_trans_map[lang] = per_lang_map
        all_sheet_trans_map = all_sheet_trans_map
            

    wb.close()
    return all_sheet_trans_map

# 读取每个sheet的翻译
def read_sheet(workbook, excen_path, sheet, platform):
    # 最大行数
    max_row = sheet.max_row + 1
    # 最大列数
    max_col = sheet.max_column + 1

    # 多语言翻译的map 例如:{zh_hans: {'trans0001': "中文"}, en: {'trans0001': "english"}}
    trans_map = OrderedDict()

    allLanguages = Project.allLanguages(project)

    # excel中key对应的列
    excel_key_col_map = OrderedDict({Platform.ios: 1, Platform.android: 2, Platform.web: 3})

    # 对应语言
    for langCol in range(1, max_col): 
        # 第一行语言描述 例如: 中文简体:zh_hans
        language_desc = sheet.cell(1, langCol).value
        if language_desc is not None:
            language_desc = str(language_desc)
            
            if Platform.ios in language_desc.lower():
                excel_key_col_map[Platform.ios] = langCol
            elif Platform.android in language_desc.lower():
                excel_key_col_map[Platform.android] = langCol
            elif Platform.web in language_desc.lower():
                excel_key_col_map[Platform.web] = langCol

            # 拆分语言的类型 
            language = language_desc
            _split_language = language_desc.split(':')
            if len(_split_language) > 0:
                language = _split_language[0]

                not_contains = True
                for lang in allLanguages:
                    if lang.excel_lang_id == language:
                        not_contains = False

                if not_contains:
                    continue

                # 翻译表根据语言初始化
                trans_lang_map = OrderedDict()
                # 第二行开始 第一列为多语言key
                for row in range(2, max_row):
                    excel_key_index = excel_key_col_map[platform]
                    excel_key_value = sheet.cell(row, excel_key_index).value
                    # 如果没有填key寻找默认key
                    if excel_key_value is None:
                        for platform_key in excel_key_col_map:
                            excel_key_col = excel_key_col_map[platform_key]
                            excel_key_value = sheet.cell(row, excel_key_col).value
                            if excel_key_value is not None:
                                break    
                                        
                    # excel中的key
                    _key = ''
                    if excel_key_value is not None:
                        _key = str(excel_key_value)
                     # key 去掉空格
                    _key = _key.replace(' ', '')
                    
                    # excel中的值
                    _text = ''
                    excel_value = sheet.cell(row=row, column=langCol).value
                    if excel_value is not None:
                        _text = str(excel_value)

                    trans_lang_map[_key] = _text
                trans_map[language] = trans_lang_map                    
    return trans_map

# 根据excel中的{1}转换拼接iOS 安卓 web格式化参数
def _convert_excel_args(text, format, platform):
    value = text
    # excel占位参数的map {占位参数: 位置}
    args_map = OrderedDict()
    # 过滤{1}占位的参数
    for result in re.finditer('{[\d]}', text):
        span = result.span()
        # excel中参数的占位格式
        arg_fmt = text[span[0]:span[1]]
        # 占位参数位置
        arg_index = arg_fmt.replace('{', '').replace('}', '')
        args_map[arg_fmt] = arg_index
    for key in args_map:
        arg_index = args_map[key]
        # web 下标从0开始 excel中翻译从1开始 所以默认减1
        if platform == Platform.web:
            arg_index = int(arg_index) - 1
            arg_index = max(0, arg_index)
        # d表示格式化参数的位置 将d替换成具体位置
        arg_fmt = format.replace('d', str(arg_index))
        # 替换excel中的占位参数
        value = value.replace(key, arg_fmt)

    return value

# 将iOS 安卓 web各端[img src=xx]参数转换成excel统一格式{img src=xx} 
def _covert_excel_imgs(text):
    value = text
    for result in re.findall(r'{img.*?}', text):
        img_tag = result.replace('{', '[').replace('}', ']')
        value = value.replace(result, img_tag)
    return value

# 逐行写入国际化源文件
def _writelines(trans_lines, dir, file_name):
    dir_path = f'{dir}'
    file_path = f'{dir_path}/{file_name}'
    # 创建文件夹
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    # 逐行写入源文件
    with open(file_path, 'w+', encoding='utf-8') as file:
        # print(f'writelines: {trans_lines}')
        file.writelines(trans_lines)

# 逐行写入ios国际化源文件
def _write_ios_localizable_strings(trans_dict, lproj_name):
    dir_path = f'{output_ios_lproj_path}{lproj_name}'
    trans_lines = []
    for key in trans_dict:
        #  逐行拼接源文件
        _key = key.replace(" ", "")
        value = trans_dict[key]
        text = _convert_excel_args(value, '%d$@', Platform.ios)
        _text = _covert_excel_imgs(text)
        # 添加转义符 \
        for escaped_char in Project.getEscapedCharacters(): 
            _text = _text.replace(f'{escaped_char}', f'\{escaped_char}')
        if len(_text) > 0: # 如果值不为空，写入国际化文件
            trans_string = f'"{_key}"="{_text}";\n'
            trans_lines.append(trans_string)
    _writelines(trans_lines, dir_path, 'Localizable.strings')

# 逐行写入android国际化源文件
def _write_android_xml(trans_dict, values_file_name):
    dir_path = f'{output_android_values_path}{values_file_name}'
    trans_lines = []
    xml_header = '<?xml version="1.0" encoding="utf-8" standalone="no"?>\n<resources>\n'
    trans_lines.append(xml_header)

    for key in trans_dict:
        #  逐行拼接源文件
        _key = key.replace(" ", "")
        value = trans_dict[key]
        text = _convert_excel_args(value, '%d$s', Platform.android)
        _text = _covert_excel_imgs(text)
        # 添加转义符 \
        for escaped_char in Project.getEscapedCharacters(): 
            _text = _text.replace(f'{escaped_char}', f'\{escaped_char}')
         # 将特殊字符 & 转为 &amp  ’
        _text = _text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        if len(_text) > 0: # 如果值不为空，写入国际化文件
            trans_string = f'    <string name="{_key}">{_text}</string>\n'
            trans_lines.append(trans_string)

    xml_tril = '</resources>'
    trans_lines.append(xml_tril)
    _writelines(trans_lines, dir_path, 'strings.xml')

# 逐行写入web国际化源文件
def _write_web_lang(trans_dict, js_file_name):
    dir_path = f'{output_web_js_path}'
    trans_lines = []
    json_header = 'export default {\n'
    trans_lines.append(json_header)
    for key in trans_dict:
        #  逐行拼接源文件
        _key = key.replace(" ", "")
        value = trans_dict[key]
        text = _convert_excel_args(value, '{d}', Platform.web)
        _text = _covert_excel_imgs(text)
        # 添加转义符 \
        for escaped_char in Project.getEscapedCharacters(): 
            _text = _text.replace(f'{escaped_char}', f'\{escaped_char}')
        if len(_text) > 0: # 如果值不为空，写入国际化文件
            trans_string = f'    "{_key}": "{_text}",\n'
            trans_lines.append(trans_string)

    json_tril = '}'
    trans_lines.append(json_tril)
    _writelines(trans_lines, dir_path, js_file_name)

# 将excel读出的多语言map写入android国际化源文件
def write_trans_to_file_android(trans_map):
    Util.clear_folder(output_android_values_path)
    for key in trans_map:
        # 每种语言对应的翻译map
        trans_dict = trans_map[key]
        values_file_name = key
        for info in Project.allLanguages(project):
            if info is None:
                continue
            if info.excel_lang_id == key:
                values_file_name = info.android_values_file_name
                break
        _write_android_xml(trans_dict, values_file_name)

# 将excel读出的多语言map写入ios国际化源文件
def write_trans_to_file_ios(trans_map):
    Util.clear_folder(output_ios_lproj_path)
    for key in trans_map:
        # 每种语言对应的翻译map
        trans_dict = trans_map[key]
        lproj_name = key
        for info in Project.allLanguages(project):
            if info.excel_lang_id == key:
                lproj_name = info.ios_lproj_file_name
                break
        _write_ios_localizable_strings(trans_dict, lproj_name)

# 将excel读出的多语言map写入web国际化源文件
def write_trans_to_file_web(trans_map):
    Util.clear_folder(output_web_js_path)
    for key in trans_map:
        # 每种语言对应的翻译map
        trans_dict = trans_map[key]
        js_file_name = key
        for info in Project.allLanguages(project):
            if info.excel_lang_id == key:
                js_file_name = info.web_js_file_name
                break
        _write_web_lang(trans_dict, js_file_name)

# 遍历读取file_path文件夹下的所有excel翻译文件
def filter_all_excels(file_path):
    excel_file_paths = []
    for root,dirs,files in os.walk(rf'{file_path}'): #遍历文件夹
        for f in files: #遍历文件
            if f.find('.xlsx') > 0: # 是excel文件
                dirPath = os.path.join(root, f) #拼接文件名路径
                excel_file_paths.append(dirPath)
    # print(f'excel_file_paths: {excel_file_paths}')
    # 文件名称排序
    sort_file_paths = sorted(excel_file_paths)
    return sort_file_paths

# 合并dir_path文件夹的所有excel翻译文件
def merge_all_excels(platform):
    dir_path = input_ios_excel_path
    if platform == Platform.android:
        dir_path = input_android_excel_path
    elif platform == Platform.ios:
        dir_path = input_ios_excel_path
    elif platform == Platform.web:
        dir_path = input_web_excel_path

    # 文件夹下所以excel的翻译汇总
    all_excel_trans_map = OrderedDict()
    file_paths = filter_all_excels(dir_path)
    for path in file_paths: #遍历文件夹
        per_excel_trans_map = read_excel(path, platform)
        for lang in per_excel_trans_map:
            per_lang_map = Util.safe_value(per_excel_trans_map, lang)
            if per_lang_map is None:
                continue
            all_lang_map = Util.safe_value(all_excel_trans_map, lang)
            if all_lang_map is not None:
                all_lang_map.update(per_lang_map)    
                all_excel_trans_map[lang] = all_lang_map
            else:
                all_excel_trans_map[lang] = per_lang_map
    return all_excel_trans_map

def trans_from_excel():

    print('正在转换...')

    # 合并ios目录下所有的excel翻译文件
    trans_map_ios= merge_all_excels(Platform.ios)
    # 写入ios国际化源文件
    write_trans_to_file_ios(trans_map_ios)

    # 合并android目录下所有的excel翻译文件
    trans_map_android = merge_all_excels(Platform.android)
    # 写入android国际化源文件
    write_trans_to_file_android(trans_map_android)

    # 合并web目录下所有的excel翻译文件
    trans_map_web = merge_all_excels(Platform.web)
    # 写入web国际化源文件
    write_trans_to_file_web(trans_map_web)
    
    print('excel转翻译源文件成功...')

# 转换翻译的app, 默认新的jblone
project = Project.jblone

# 函数入口 第一个参数是转换语言的app项目名称  eg: python3 trans_from_excel.py partybox / jblone
if __name__ == '__main__':
    if len(sys.argv) > 1: # 传递了参数
        project = sys.argv[1] # 第一个参数是转换语言的app项目名称 partybox / jblone
    
    trans_from_excel()